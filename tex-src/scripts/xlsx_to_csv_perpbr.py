#!/usr/bin/env python3
"""scripts/xlsx_to_csv_perpbr.py   v1.1  (2025-06-08)
────────────────────────────────────────────────────────
CHANGELOG:
- 2025-06-08  v1.1 : フォーミュラセルの数値読み取りに対応
- 2025-06-08  v1.0 : 初版 — perpbr.xlsx 一括 CSV 変換
"""

from __future__ import annotations

from pathlib import Path
import argparse
import pandas as pd
import openpyxl

# ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
XLSX_DIR = ROOT / "tex-src" / "data" / "earn" / "perpbr"
OUT_CSV = ROOT / "tex-src" / "data" / "earn" / "perpbr.csv"


def read_xlsx(path: Path) -> pd.DataFrame:
    """Read PER/PBR excel and normalize formula cells."""

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    header = []
    for idx in range(1, 15):
        val = ws.cell(row=4, column=idx).value
        header.append(val if val is not None else f"Unnamed: {idx - 1}")
    rows = []
    for r in range(5, ws.max_row + 1):
        row: list[float | str | None] = []
        for c in range(1, 15):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.startswith("="):
                try:
                    val = float(val.lstrip("="))
                except ValueError:
                    pass
            row.append(val)
        rows.append(row)
    return pd.DataFrame(rows, columns=header)


def main() -> None:
    argparse.ArgumentParser(description="perpbr xlsx to single CSV").parse_args()

    frames: list[pd.DataFrame] = []
    existing_months: set[str] = set()

    if OUT_CSV.exists():
        existing = pd.read_csv(OUT_CSV)
        frames.append(existing)
        if "Year/Month" in existing.columns:
            existing_months.update(existing["Year/Month"].astype(str).unique())

    for xlsx in sorted(XLSX_DIR.glob("perpbr*.xlsx")):
        df = read_xlsx(xlsx)
        month = str(df["Year/Month"].iloc[0])
        if month in existing_months:
            continue
        frames.append(df)

    if not frames:
        print("🟡 更新対象なし")
        return

    merged = pd.concat(frames, ignore_index=True)
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    merged.to_csv(OUT_CSV, index=False, encoding="utf-8")
    print(f"✅ 出力: {OUT_CSV}")


if __name__ == "__main__":
    main()
